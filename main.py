import random
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

def ran():
    for i in range(0,5):
                a[i] = random.random()
def func():
    def calc():
        a1 = (a[0] + a[1] )/2
        a2 = (a[2] + a[3] )/2
        a3 = (a[4] + a[5] )/2

        big_a = (a1+a2+a3)/3

        k1 = big_a - a1
        k2 = big_a - a2
        k3 = big_a - a3

        if k1 < 0:
            k1 = -1*k1
        if k2 < 0:
            k2 = -1*k2
        if k3 < 0:
            k3 = -1*k3

        big_k  = (k1+k2+k3)/3
        per = (big_k/big_a)*100
        

    calc()
    m,b = 0,0
    number = 0
    for m in range(1,20):
        for b in range(1,7):
            global per
            list = f'{letter[b]}{m}'
            number = a[b-1]
            
            ws[list] = number
            #ws[f'G{m}'] = per 
            ran()
            calc()



x1,x2,y1,y2,z1,z2 = 0,0,0,0,0,0
a  = [x1,x2,y1,y2,z1,z2]
letter = ['A','B','C','D','E','F','G']

for _ in range(0,10):
    
    ran()
    func()

    _ += 1

wb.save("sample2.xlsx")
print('done')
